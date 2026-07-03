"""Excel 导出服务测试。

验证 build_inventory_excel 生成的工作簿结构与数量聚合正确。
"""
import os
import sys
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from services.export import build_inventory_excel


def _row(**kw):
    """构造一条 inventory 记录 dict（模拟 sqlite3.Row 的字段访问）"""
    base = dict(
        package_type="QFP", spec="14x20", plating_zone="不镀银",
        surface_treatment="粗化", manufacturer="AAMI", batch_no="LOT1",
        production_date="2024-01-01", expiry_date="2025-01-01",
        quantity="5", note="",
    )
    base.update(kw)
    return base


class TestBuildInventoryExcel(unittest.TestCase):
    def test_two_sheets_present(self):
        buf = build_inventory_excel([_row()])
        wb = openpyxl.load_workbook(buf)
        self.assertIn("库存汇总", wb.sheetnames)
        self.assertIn("库存明细", wb.sheetnames)

    def test_detail_sheet_has_all_rows(self):
        rows = [_row(batch_no="LOT1"), _row(batch_no="LOT2"), _row(batch_no="LOT3")]
        buf = build_inventory_excel(rows)
        wb = openpyxl.load_workbook(buf)
        ws = wb["库存明细"]
        # 表头 1 行 + 3 条记录
        self.assertEqual(ws.max_row, 1 + 3)

    def test_summary_sheet_aggregates_by_group_key(self):
        # 同组（含不同批号）3 条，数量 5+5+3 = 13；另一组 1 条
        rows = [
            _row(batch_no="LOT1", quantity="5"),
            _row(batch_no="LOT2", quantity="5"),
            _row(batch_no="LOT3", quantity="3"),
            _row(batch_no="LOT1", manufacturer="华天", quantity="2"),
        ]
        buf = build_inventory_excel(rows)
        wb = openpyxl.load_workbook(buf)
        ws = wb["库存汇总"]
        # 表头 + 2 个分组
        self.assertEqual(ws.max_row, 1 + 2)
        # 找到 AAMI 分组行，验证聚合数量与批次数
        summary = {}
        for r in ws.iter_rows(min_row=2, values_only=True):
            pt, sp, pz, st, mf, qty, count = r
            summary[mf] = (qty, count)
        self.assertEqual(summary["AAMI"], ("13", 3))
        self.assertEqual(summary["华天"], ("2", 1))

    def test_empty_rows_still_produces_valid_workbook(self):
        buf = build_inventory_excel([])
        wb = openpyxl.load_workbook(buf)
        ws = wb["库存明细"]
        self.assertEqual(ws.max_row, 1)  # 只有表头


if __name__ == "__main__":
    unittest.main()
